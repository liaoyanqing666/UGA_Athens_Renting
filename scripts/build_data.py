from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "data"

ZH_HEADERS = [
    "名称",
    "地理位置",
    "位置",
    "房型",
    "大小",
    "人均价格(最低)",
    "家具",
    "距离Boyd开车(mile/min)",
    "骑车(mile/min)",
    "公交/shuttle bus",
    "谷歌评分",
    "谷歌评价",
    "各种费用(如有写明)(不含申请费)",
    "社区设施(运动)",
    "注意事项/备注",
    "网址",
]

EN_HEADERS = [
    "Name",
    "Region",
    "Location",
    "Layouts",
    "Size",
    "Min Price Per Person",
    "Furnishing",
    "Drive to Boyd (mile/min)",
    "Bike (mile/min)",
    "Bus/shuttle",
    "Google Rating",
    "Google Reviews",
    "Fees (if disclosed, excluding application fee)",
    "Community Amenities (Sports)",
    "Notes",
    "Website",
]

DATASETS = [
    {
        "source": ROOT / "26fall renting.xlsx",
        "output": OUTPUT_DIR / "apartments.json",
        "headers": ZH_HEADERS,
        "fields": {
            "name": "名称",
            "region": "地理位置",
            "location": "位置",
            "layouts": "房型",
            "sizes": "大小",
            "prices": "人均价格(最低)",
            "furnished": "家具",
            "drive": "距离Boyd开车(mile/min)",
            "bike": "骑车(mile/min)",
            "bus": "公交/shuttle bus",
            "rating": "谷歌评分",
            "reviews": "谷歌评价",
            "fees": "各种费用(如有写明)(不含申请费)",
            "amenities": "社区设施(运动)",
            "notes": "注意事项/备注",
            "website": "网址",
        },
    },
    {
        "source": ROOT / "26fall renting EN.xlsx",
        "output": OUTPUT_DIR / "apartments_EN.json",
        "headers": EN_HEADERS,
        "fields": {
            "name": "Name",
            "region": "Region",
            "location": "Location",
            "layouts": "Layouts",
            "sizes": "Size",
            "prices": "Min Price Per Person",
            "furnished": "Furnishing",
            "drive": "Drive to Boyd (mile/min)",
            "bike": "Bike (mile/min)",
            "bus": "Bus/shuttle",
            "rating": "Google Rating",
            "reviews": "Google Reviews",
            "fees": "Fees (if disclosed, excluding application fee)",
            "amenities": "Community Amenities (Sports)",
            "notes": "Notes",
            "website": "Website",
        },
    },
]


@dataclass
class UnitOption:
    layout: str
    bedrooms: float | None
    bathrooms: float | None
    size_sqft: float | None
    min_price_per_person: float | None
    raw_size: str | None
    raw_price: str | None


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if int(value) == value:
            return str(int(value))
        return str(value)
    return " ".join(str(value).replace("\n", " ").split())


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", normalized.lower()).strip("-")
    return slug or "apartment"


def parse_number(value: str) -> float | None:
    match = re.search(r"\d+(?:\.\d+)?", value)
    return float(match.group()) if match else None


def parse_layout(layout: str) -> tuple[float | None, float | None]:
    layout = layout.strip().lower()
    if layout == "studio":
        return 0.0, 1.0
    match = re.fullmatch(r"(\d+(?:\.\d+)?)b(\d+(?:\.\d+)?)b", layout)
    if not match:
        return None, None
    return float(match.group(1)), float(match.group(2))


def parse_token_values(raw: Any) -> list[str]:
    text = normalize_space(raw)
    if not text or text == "未知":
        return []
    return re.findall(r"\S+", text)


def parse_commute_pair(raw: Any) -> dict[str, float | None | str]:
    text = normalize_space(raw)
    match = re.search(r"(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)", text)
    if not match:
        return {"raw": text, "miles": None, "minutes": None}
    return {
        "raw": text,
        "miles": float(match.group(1)),
        "minutes": float(match.group(2)),
    }


def parse_bus(raw: Any) -> dict[str, Any]:
    text = normalize_space(raw)
    minutes = parse_number(text)
    lowered = text.lower()
    has_shuttle = False if "无shuttle" in lowered else ("shuttle" in lowered if lowered else None)
    return {"raw": text, "minutes": minutes, "hasShuttle": has_shuttle}


def parse_amenities(raw: Any) -> list[str]:
    text = normalize_space(raw)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[，、；;/]+", text) if item.strip()]


def build_units(layouts_raw: Any, sizes_raw: Any, prices_raw: Any) -> list[UnitOption]:
    layouts = parse_token_values(layouts_raw)
    sizes = parse_token_values(sizes_raw)
    prices = parse_token_values(prices_raw)

    units: list[UnitOption] = []
    for index, layout in enumerate(layouts):
        size_token = sizes[index] if index < len(sizes) else None
        price_token = prices[index] if index < len(prices) else None
        bedrooms, bathrooms = parse_layout(layout)
        units.append(
            UnitOption(
                layout=layout,
                bedrooms=bedrooms,
                bathrooms=bathrooms,
                size_sqft=parse_number(size_token) if size_token else None,
                min_price_per_person=parse_number(price_token) if price_token else None,
                raw_size=size_token,
                raw_price=price_token,
            )
        )
    return units


def build_dataset(config: dict[str, Any]) -> None:
    source_file = config["source"]
    output_file = config["output"]
    headers = config["headers"]
    fields = config["fields"]

    if not source_file.exists():
        raise SystemExit(f"Missing source Excel file: {source_file.name}")

    workbook = load_workbook(source_file, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]

    notes = dict(zip(headers, rows[1]))
    apartments = []
    seen_slugs: dict[str, int] = {}

    for row_index, row in enumerate(rows[2:], start=3):
        if not any(value is not None for value in row):
            continue

        record = dict(zip(headers, row[: len(headers)]))
        name = normalize_space(record[fields["name"]])
        base_slug = slugify(name)
        seen_slugs[base_slug] = seen_slugs.get(base_slug, 0) + 1
        slug = base_slug if seen_slugs[base_slug] == 1 else f"{base_slug}-{seen_slugs[base_slug]}"

        units = build_units(record[fields["layouts"]], record[fields["sizes"]], record[fields["prices"]])
        prices = [unit.min_price_per_person for unit in units if unit.min_price_per_person is not None]
        sizes = [unit.size_sqft for unit in units if unit.size_sqft is not None]
        drive = parse_commute_pair(record[fields["drive"]])
        bike = parse_commute_pair(record[fields["bike"]])
        bus = parse_bus(record[fields["bus"]])

        apartments.append(
            {
                "id": slug,
                "rowNumber": row_index,
                "name": name,
                "region": normalize_space(record[fields["region"]]),
                "location": normalize_space(record[fields["location"]]),
                "furnished": normalize_space(record[fields["furnished"]]),
                "googleRating": parse_number(normalize_space(record[fields["rating"]])),
                "googleReviewSummary": normalize_space(record[fields["reviews"]]),
                "fees": normalize_space(record[fields["fees"]]),
                "amenitiesText": normalize_space(record[fields["amenities"]]),
                "amenities": parse_amenities(record[fields["amenities"]]),
                "notes": normalize_space(record[fields["notes"]]),
                "website": normalize_space(record[fields["website"]]),
                "commute": {
                    "drive": drive,
                    "bike": bike,
                    "bus": bus,
                },
                "units": [unit.__dict__ for unit in units],
                "summary": {
                    "unitCount": len(units),
                    "minPricePerPerson": min(prices) if prices else None,
                    "maxPricePerPerson": max(prices) if prices else None,
                    "minSizeSqft": min(sizes) if sizes else None,
                    "maxSizeSqft": max(sizes) if sizes else None,
                },
                "raw": {
                    "layouts": normalize_space(record[fields["layouts"]]),
                    "sizes": normalize_space(record[fields["sizes"]]),
                    "prices": normalize_space(record[fields["prices"]]),
                    "drive": normalize_space(record[fields["drive"]]),
                    "bike": normalize_space(record[fields["bike"]]),
                    "bus": normalize_space(record[fields["bus"]]),
                },
            }
        )

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFile": source_file.name,
        "sheetName": worksheet.title,
        "columns": headers,
        "columnNotes": notes,
        "recordCount": len(apartments),
        "apartments": apartments,
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {output_file}")


def main() -> None:
    for dataset in DATASETS:
        build_dataset(dataset)


if __name__ == "__main__":
    main()
